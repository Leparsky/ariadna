#https://habr.com/post/322608/
#https://habr.com/post/250921/
from selenium import webdriver
import selenium.common.exceptions
running = True
from selenium.webdriver.chrome.options import DesiredCapabilities
from selenium.webdriver.common.proxy import Proxy, ProxyType

from openpyxl import Workbook

BASE_URL = 'https://www.ariadna-96.ru'     #адрес сайта для парсинга




#https://gist.github.com/tushortz/cba8b25f9d80f584f807b65890f37be5
def get_proxies():
    co = webdriver.ChromeOptions()
    co.add_argument("log-level=3")
    co.add_argument("--headless")
    driver = webdriver.Chrome(chrome_options=co)
    driver.get("https://free-proxy-list.net/")

    PROXIES = []
    proxies = driver.find_elements_by_css_selector("tr[role='row']")
    for p in proxies:
        result = p.text.split(" ")

        if result[-1] == "yes":
            PROXIES.append(result[0]+":"+result[1])

    driver.close()
    return PROXIES





def proxy_driver(PROXIES):
    co = webdriver.ChromeOptions()
    co.add_argument("log-level=3")
    #co.add_argument("--headless")
    prox = Proxy()

    if PROXIES:
        pxy = PROXIES[-1]
    else:
        print("--- Proxies used up (%s)" % len(PROXIES))
        PROXIES = get_proxies()

    prox.proxy_type = ProxyType.MANUAL
    prox.http_proxy = pxy
    prox.socks_proxy = pxy
    prox.ssl_proxy = pxy

    capabilities = webdriver.DesiredCapabilities.CHROME
    prox.add_to_capabilities(capabilities)

    driver = webdriver.Chrome(chrome_options=co, desired_capabilities=capabilities)

    return driver
def get_page(url,driver):
    running = True
    while running:
        try:
            driver.get(url)

            assert "Телефоны"  in driver.page_source
            running = False
            #running = True
        except:
            new = ALL_PROXIES.pop()
            # reassign driver if fail to switch proxy
            driver = proxy_driver(ALL_PROXIES)
            print("--- Switched proxy to: %s" % new)
            time.sleep(1)

def main():
    #ALL_PROXIES = get_proxies()
    #driver = proxy_driver(ALL_PROXIES)
    co = webdriver.ChromeOptions()
    co.add_argument("log-level=3")
    # co.add_argument("--headless")
    driver = webdriver.Chrome(chrome_options=co)

    driver.maximize_window()
    lstCatPages = []
    lstGoodPages = []
    # страницы каталога
    driver.get(BASE_URL+"/user/login")
    element = driver.find_element_by_xpath("//input[@id='login-username']")
    element.send_keys("nata.vityaz@mail.ru")
    element = driver.find_element_by_xpath("//input[@id='login-password']")
    element.send_keys("q1e3t5u7")
    element = driver.find_element_by_xpath("//button[@id='postbut']")
    element.click()

    for element in driver.find_elements_by_xpath("//a[@class='leftside_inner__content___menu____list_____item-link']"):
        lstCatPages.append(element.get_attribute("href"))
    for cat in lstCatPages:
        driver.get(cat)
        for element in driver.find_elements_by_xpath("//a[@class='item_info__title-link']"):
            lstGoodPages.append(element.get_attribute("href"))

    #lstGoodPages.append("https://www.ariadna-96.ru/detskaya-odezhda/gnkmain-kostyumyi/1528")
    #stGoodPages.append("https://www.ariadna-96.ru/detskaya-odezhda/gnkbaby-kostyumyi/1526")
    wb = Workbook()
    ws = wb.active
    pagesproxycount = 30
    ws.cell(column=1+0, row=1, value="ID")
    ws.cell(column=1+1, row=1, value="Название")
    ws.cell(column=1+2, row=1, value="Оригинальное")
    ws.cell(column=1+3, row=1, value="название")
    ws.cell(column=1+4, row=1, value="Цена")
    ws.cell(column=1+5, row=1, value="Количество")
    ws.cell(column=1+6, row=1, value="Размер")
    ws.cell(column=1+7, row=1, value="Артикул")
    ws.cell(column=1+8, row=1, value="Цвет")
    ws.cell(column=1+9, row=1, value="Единица измерения")
    ws.cell(column=1+10, row=1, value="Описание")
    ws.cell(column=1+11, row=1, value="Фотография")
    ws.cell(column=1+12, row=1, value="Альбом")
    ws.cell(column=1+13, row=1, value="Позиция")
    ws.cell(column=1+14, row=1, value="Отображать комментарий")
    ws.cell(column=1+15, row=1, value="Отображать в каталоге")
    ws.cell(column=1+16, row=1, value="Включить ряды")
    ws.cell(column=1+17, row=1, value="Ссылка на источник")
    ws.cell(column=1+18, row=1, value="Обновить фото")

    i = 1
    j = 1

    for page in lstGoodPages:
        url = ""  # url
        cost = ""  # Цена
        size = ""  # Размер
        art = ""  # Артикул
        color = ""  # Цвет
        edizm = ""  # Единица    измерения
        descr = ""  # // Описание
        album = ""  # // Альбом
        position = ""  # Позиция
        picture = ""  # // Ссылка на картинку
        name = ""  # Название
        driver.get(page)
        j += 1
        print("товар {} из {}".format(j-1,len(lstGoodPages)))
        print(page)
        dictCS = {}
        url = page

        cost = driver.find_element_by_xpath("//div[@class='i-price bold']").text[:-4].strip()
        name = driver.find_element_by_xpath("//div[@class='pos_info col-sm-7']/h1").text
        art = driver.find_element_by_xpath("//li[@class='pos_info__tab___har-item'][1]/span[@class='color-802e80 bold']").text
        size = driver.find_element_by_xpath("//li[@class='pos_info__tab___har-item'][2]/span[@class='color-802e80 bold']").text
        descr = driver.find_element_by_xpath("//div[@class='pos_info__tab tab-content']").text
        driver.find_element_by_xpath("//li[@id='itemdescr_tab']/a[@class='pos_info__menu___item-link']").click()
        descr += driver.find_element_by_xpath("//div[@id='itemdescr']").text
        try:
            driver.find_element_by_xpath("//li[@class='indicators_item'][0]").click()
        except selenium.common.exceptions.NoSuchElementException:
            None
        picture = driver.find_element_by_xpath("//div[@class='slides_item item active']/a").get_attribute("href")
        album=driver.find_element_by_xpath("//ol[@class='crumbs']").text.split("одежда")[1]

        for CS in driver.find_elements_by_xpath("//div[@class='i-har row pos-rel']"):
            a = CS.find_element_by_xpath(".//div[@class='color_title']").text
            b = ""
            for CS1 in CS.find_elements_by_xpath(".//select[@class='form-control itemsizes']/option"):
                b+=CS1.text.split("(")[0]+","
            dictCS[a]=b[:-1]
        if not len(dictCS):
            continue
        #try:
        #    element = driver.find_element_by_xpath("//li[@class='b-product-data__item b-product-data__item_type_available']")
        #    descr = (element.text)+' , '
        #except selenium.common.exceptions.NoSuchElementException:
        #    None
        for cs in dictCS.keys():
            i += 1
            ws.cell(column=1+1, row=i, value=name)
            ws.cell(column=1+4, row=i, value=cost)
            ws.cell(column=1+6, row=i, value=dictCS[cs])
            ws.cell(column=1+7, row=i, value=art)
            ws.cell(column=1+8, row=i, value=cs)
            ws.cell(column=1+9, row=i, value=edizm)
            ws.cell(column=1+10,row=i, value=descr)
            ws.cell(column=1+12,row=i, value=album)
            ws.cell(column=1+13,row=i, value=i - 1)
            ws.cell(column=1+17,row=i, value=picture)
            ws.cell(column=1+18,row=i, value=url)
    wb.save(filename="c:\\tmp\\ariadna.xlsx")


main()
#запуск приложения
